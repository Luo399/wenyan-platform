#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel读取工具模块

提供Excel文件读取、表头解析、数据提取等功能
遵循单一职责原则，专注于Excel文件的读取操作
"""

import os
import logging
from typing import List, Dict, Any, Optional, Callable

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError("请先安装 openpyxl: pip install openpyxl")

logger = logging.getLogger(__name__)

class ExcelHeader:
    """
    Excel表头信息类
    
    属性:
        chinese_names: 第1行中文列名列表
        english_names: 第2行英文属性名列表
    """
    
    def __init__(self, chinese_names: List[str], english_names: List[str]):
        self.chinese_names = chinese_names
        self.english_names = english_names
    
    def __repr__(self) -> str:
        return f"ExcelHeader(chinese_names={self.chinese_names[:5]}..., english_names={self.english_names[:5]}...)"
    
    @property
    def valid_columns(self) -> List[str]:
        """获取有效的英文列名列表（过滤空值）"""
        return [name for name in self.english_names if name]

class ExcelConfig:
    """
    Excel读取配置类
    
    参数:
        input_file: 输入Excel文件路径
        sheet_name: 工作表名称
        header_row: 中文列名所在行（从1开始）
        property_row: 英文属性名所在行（从1开始）
        data_start_row: 数据起始行（从1开始）
        empty_value_replacement: 空值替换值，默认为None
        validate_file: 是否在初始化时验证文件存在，默认为False
    """
    
    def __init__(
        self,
        input_file: str,
        sheet_name: str = None,
        header_row: int = 1,
        property_row: int = 2,
        data_start_row: int = 3,
        empty_value_replacement: Any = None,
        validate_file: bool = False
    ):
        self.input_file = input_file
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.property_row = property_row
        self.data_start_row = data_start_row
        self.empty_value_replacement = empty_value_replacement
        
        # 可选：验证输入文件
        if validate_file and not os.path.exists(input_file):
            raise FileNotFoundError(f"输入文件不存在: {input_file}")

def open_workbook(file_path: str, read_only: bool = True) -> Workbook:
    """
    打开Excel工作簿
    
    参数:
        file_path: Excel文件路径
        read_only: 是否以只读模式打开
    
    返回:
        Workbook对象
    
    异常:
        FileNotFoundError: 文件不存在
        ValueError: 文件格式不正确
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    try:
        return openpyxl.load_workbook(file_path, read_only=read_only, data_only=True)
    except Exception as e:
        raise ValueError(f"无法打开Excel文件: {str(e)}")

def _clean_value(value: Any, replacement: Any = None) -> Any:
    """
    清理单元格值
    
    参数:
        value: 原始值
        replacement: 空值替换值
    
    返回:
        清理后的值
    """
    if value is None:
        return replacement
    
    if isinstance(value, str):
        cleaned = value.strip()
        
        # 处理特殊空值标记
        if cleaned.lower() in ['', 'none', 'null', 'nan', 'n/a', '-', '—']:
            return replacement
        
        # 仅当字符串完全由0-9数字组成时才转换为整数
        if cleaned.isdigit() and all(c in '0123456789' for c in cleaned):
            return int(cleaned)
        
        # 尝试转换为浮点数
        try:
            return float(cleaned)
        except ValueError:
            pass
        
        return cleaned
    
    # 对于其他类型（数字、日期等），保持原样
    return value

def read_sheet_headers(
    sheet: Worksheet,
    header_row: int = 1,
    property_row: int = 2
) -> ExcelHeader:
    """
    读取工作表表头信息
    
    参数:
        sheet: Worksheet对象
        header_row: 中文列名所在行（从1开始）
        property_row: 英文属性名所在行（从1开始）
    
    返回:
        ExcelHeader对象，包含中文和英文列名
    """
    chinese_names: List[str] = []
    english_names: List[str] = []
    
    for col in range(1, sheet.max_column + 1):
        chinese_val = sheet.cell(row=header_row, column=col).value
        english_val = sheet.cell(row=property_row, column=col).value
        
        chinese_names.append(_clean_value(chinese_val))
        english_names.append(_clean_value(english_val))
    
    return ExcelHeader(chinese_names, english_names)

def read_sheet_data(
    sheet: Worksheet,
    header: ExcelHeader,
    data_start_row: int = 3,
    empty_value_replacement: Any = None,
    filter_func: Optional[Callable[[Dict[str, Any]], bool]] = None
) -> List[Dict[str, Any]]:
    """
    读取工作表数据
    
    参数:
        sheet: Worksheet对象
        header: ExcelHeader对象
        data_start_row: 数据起始行（从1开始）
        empty_value_replacement: 空值替换值
        filter_func: 行过滤函数，返回True保留该行
    
    返回:
        字典列表，每个字典代表一行数据
    """
    data: List[Dict[str, Any]] = []
    
    for row_num in range(data_start_row, sheet.max_row + 1):
        row_data: Dict[str, Any] = {}
        
        for col_idx, (chinese_name, english_name) in enumerate(
            zip(header.chinese_names, header.english_names)
        ):
            if not english_name:
                continue
            
            cell_value = sheet.cell(row=row_num, column=col_idx + 1).value
            row_data[english_name] = _clean_value(
                cell_value, 
                replacement=empty_value_replacement
            )
        
        # 检查是否为空行
        if not any(v is not None for v in row_data.values()):
            continue
        
        # 应用过滤函数
        if filter_func and not filter_func(row_data):
            continue
        
        data.append(row_data)
    
    return data

def read_sheet(
    workbook: Workbook,
    sheet_name: str,
    config: ExcelConfig
) -> ExcelHeader:
    """
    读取整个工作表的表头信息（用于快速检查）
    
    参数:
        workbook: Workbook对象
        sheet_name: 工作表名称
        config: ExcelConfig配置对象
    
    返回:
        ExcelHeader对象
    """
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表不存在: {sheet_name}")
    
    sheet = workbook[sheet_name]
    return read_sheet_headers(sheet, config.header_row, config.property_row)

def read_sheet_full(
    workbook: Workbook,
    sheet_name: str,
    config: ExcelConfig
) -> tuple:
    """
    读取整个工作表的完整数据
    
    参数:
        workbook: Workbook对象
        sheet_name: 工作表名称
        config: ExcelConfig配置对象
    
    返回:
        (header, rows) 元组，header为ExcelHeader对象，rows为数据行列表
    """
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表不存在: {sheet_name}")
    
    sheet = workbook[sheet_name]
    header = read_sheet_headers(sheet, config.header_row, config.property_row)
    rows = read_sheet_data(
        sheet, 
        header, 
        config.data_start_row,
        config.empty_value_replacement
    )
    
    logger.info(f"读取工作表 '{sheet_name}': {len(rows)} 行数据")
    return (header, rows)

def get_english_sheets(workbook: Workbook) -> List[str]:
    """
    获取所有英文命名的工作表名称
    
    参数:
        workbook: Workbook对象
    
    返回:
        英文命名的工作表名称列表
    """
    return [name for name in workbook.sheetnames if name and name[0].isascii()]

def print_sheet_info(file_path: str) -> None:
    """
    打印Excel文件信息（用于调试）
    
    参数:
        file_path: Excel文件路径
    """
    try:
        workbook = open_workbook(file_path)
        
        print("=" * 60)
        print(f"Excel文件: {file_path}")
        print("=" * 60)
        print(f"工作表数量: {len(workbook.sheetnames)}")
        print(f"工作表列表: {workbook.sheetnames}")
        print("-" * 60)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            header = read_sheet_headers(sheet)
            
            print(f"\n工作表: '{sheet_name}'")
            print(f"  行数: {sheet.max_row}, 列数: {sheet.max_column}")
            print(f"  中文列名: {header.chinese_names}")
            print(f"  英文列名: {header.english_names}")
        
        print("\n" + "=" * 60)
        
    except Exception as e:
        logger.error(f"读取文件信息失败: {str(e)}")