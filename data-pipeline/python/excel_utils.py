#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import json
import logging

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError("请先安装 openpyxl: pip install openpyxl")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

class ExcelHeader:
    def __init__(self, chinese_names, english_names):
        self.chinese_names = chinese_names
        self.english_names = english_names

class ExcelConfig:
    def __init__(self, input_file, output_dir, sheet_name=None, header_row=1, 
                 property_row=2, data_start_row=3, encoding='utf-8', 
                 json_indent=2, empty_value_replacement=None):
        self.input_file = input_file
        self.output_dir = output_dir
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.property_row = property_row
        self.data_start_row = data_start_row
        self.encoding = encoding
        self.json_indent = json_indent
        self.empty_value_replacement = empty_value_replacement
        
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"输入文件不存在: {input_file}")
        os.makedirs(output_dir, exist_ok=True)
    
    def get_output_path(self, filename):
        return os.path.join(self.output_dir, filename)

def open_workbook(file_path, read_only=True):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    try:
        return openpyxl.load_workbook(file_path, read_only=read_only, data_only=True)
    except Exception as e:
        raise ValueError(f"无法打开Excel文件: {str(e)}")

def _clean_value(value, replacement=None):
    if value is None:
        return replacement
    if isinstance(value, str):
        cleaned = value.strip()
        if cleaned.lower() in ['', 'none', 'null', 'nan', 'n/a', '-', '—']:
            return replacement
        # 仅当字符串完全由0-9数字组成时才转换为整数
        # 排除带圆圈的数字等特殊字符
        if cleaned.isdigit() and all(c in '0123456789' for c in cleaned):
            return int(cleaned)
        try:
            return float(cleaned)
        except ValueError:
            pass
        return cleaned
    return value

def read_sheet_headers(sheet, header_row=1, property_row=2):
    chinese_names = []
    english_names = []
    for col in range(1, sheet.max_column + 1):
        chinese_val = sheet.cell(row=header_row, column=col).value
        english_val = sheet.cell(row=property_row, column=col).value
        chinese_names.append(_clean_value(chinese_val))
        english_names.append(_clean_value(english_val))
    return ExcelHeader(chinese_names, english_names)

def read_sheet_data(sheet, header, data_start_row=3, empty_value_replacement=None, filter_func=None):
    data = []
    for row_num in range(data_start_row, sheet.max_row + 1):
        row_data = {}
        for col_idx, (chinese_name, english_name) in enumerate(
            zip(header.chinese_names, header.english_names)
        ):
            if not english_name:
                continue
            cell_value = sheet.cell(row=row_num, column=col_idx + 1).value
            row_data[english_name] = _clean_value(cell_value, replacement=empty_value_replacement)
        
        if not any(v is not None for v in row_data.values()):
            continue
        
        if filter_func and not filter_func(row_data):
            continue
        
        data.append(row_data)
    return data

def read_sheet(workbook, sheet_name, config):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表不存在: {sheet_name}")
    sheet = workbook[sheet_name]
    header = read_sheet_headers(sheet, config.header_row, config.property_row)
    rows = read_sheet_data(sheet, header, config.data_start_row, config.empty_value_replacement)
    logger.info(f"读取工作表 '{sheet_name}': {len(rows)} 行数据")
    return header, rows

def convert_letter_to_index(letter):
    if not letter:
        return None
    letter = str(letter).strip().upper()
    if letter in ['A', 'B', 'C', 'D']:
        return ord(letter) - ord('A')
    return None

def group_by_column(data, column_name):
    groups = {}
    for row in data:
        key = str(row.get(column_name, '')).strip()
        if not key:
            continue
        if key not in groups:
            groups[key] = []
        groups[key].append(row)
    return groups

def save_json(data, file_path, encoding='utf-8', indent=2):
    try:
        with open(file_path, 'w', encoding=encoding) as f:
            json.dump(data, f, ensure_ascii=False, indent=indent)
        logger.info(f"已保存: {file_path}")
    except Exception as e:
        raise IOError(f"保存文件失败: {str(e)}")

def excel_to_json(input_file, output_dir, sheet_name=None, transform_func=None, **config_kwargs):
    logger.info(f"开始转换 Excel -> JSON")
    logger.info(f"输入文件: {input_file}")
    logger.info(f"输出目录: {output_dir}")
    
    try:
        workbook = open_workbook(input_file)
        
        if sheet_name:
            config = ExcelConfig(input_file, output_dir, sheet_name, **config_kwargs)
            header, rows = read_sheet(workbook, sheet_name, config)
            
            if transform_func:
                result = transform_func(rows)
            else:
                result = rows
            
            output_path = config.get_output_path(f"{sheet_name}.json")
            save_json(result, output_path, encoding=config.encoding, indent=config.json_indent)
            logger.info(f"成功转换工作表: {sheet_name}")
            return result
        
        else:
            processed_count = 0
            for name in workbook.sheetnames:
                if not name or not name[0].isascii():
                    continue
                
                try:
                    config = ExcelConfig(input_file, output_dir, name, **config_kwargs)
                    header, rows = read_sheet(workbook, name, config)
                    
                    if transform_func:
                        result = transform_func(rows)
                    else:
                        result = rows
                    
                    output_path = config.get_output_path(f"{name}.json")
                    save_json(result, output_path, encoding=config.encoding, indent=config.json_indent)
                    processed_count += 1
                    logger.info(f"成功转换工作表: {name}")
                
                except Exception as e:
                    logger.error(f"转换工作表 '{name}' 失败: {str(e)}")
            
            logger.info(f"转换完成！共处理 {processed_count} 个工作表")
            return None
    
    except Exception as e:
        logger.error(f"转换失败: {str(e)}")
        raise

def print_sheet_info(file_path):
    try:
        workbook = open_workbook(file_path)
        print("=" * 60)
        print(f"Excel文件: {file_path}")
        print("=" * 60)
        print(f"工作表数量: {len(workbook.sheetnames)}")
        print(f"工作表列表: {workbook.sheetnames}")
        print("-" * 60)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            header = read_sheet_headers(sheet)
            print(f"\n工作表: '{sheet_name}'")
            print(f"  行数: {sheet.max_row}, 列数: {sheet.max_column}")
            print(f"  中文列名: {header.chinese_names}")
            print(f"  英文列名: {header.english_names}")
        
        print("\n" + "=" * 60)
        
    except Exception as e:
        logger.error(f"读取文件信息失败: {str(e)}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Excel转JSON通用工具")
    parser.add_argument("input", help="输入Excel文件路径")
    parser.add_argument("-o", "--output", default="./output", help="输出目录")
    parser.add_argument("-s", "--sheet", help="指定工作表名称")
    parser.add_argument("-l", "--list", action="store_true", help="列出工作表信息")
    
    args = parser.parse_args()
    
    if args.list:
        print_sheet_info(args.input)
    else:
        excel_to_json(input_file=args.input, output_dir=args.output, sheet_name=args.sheet)