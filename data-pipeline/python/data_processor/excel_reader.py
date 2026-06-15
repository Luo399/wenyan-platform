#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel读取模块
负责从Excel文件中读取原始数据
"""

import os
import logging
from typing import Dict, List, Any, Optional

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.workbook.workbook import Workbook
except ImportError:
    raise ImportError("请先安装 openpyxl: pip install openpyxl")

logger = logging.getLogger(__name__)


class ExcelHeader:
    """表头信息"""
    def __init__(self, chinese_names: List[str], english_names: List[str]):
        self.chinese_names = chinese_names
        self.english_names = english_names
    
    def __repr__(self):
        return f"ExcelHeader(chinese_names={self.chinese_names}, english_names={self.english_names})"


class ExcelReader:
    """Excel文件读取器"""
    
    def __init__(self, file_path: str, read_only: bool = True):
        """
        初始化Excel读取器
        
        :param file_path: Excel文件路径
        :param read_only: 是否以只读模式打开
        """
        self.file_path = file_path
        self.read_only = read_only
        self.workbook: Optional[Workbook] = None
        self._validate_file()
    
    def _validate_file(self):
        """验证文件是否存在"""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel文件不存在: {self.file_path}")
        
        if not self.file_path.lower().endswith(('.xlsx', '.xlsm')):
            raise ValueError(f"不支持的文件格式: {self.file_path}")
    
    def open_workbook(self):
        """打开工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(
                self.file_path,
                read_only=self.read_only,
                data_only=True
            )
            logger.info(f"成功打开Excel文件: {self.file_path}")
        except Exception as e:
            raise RuntimeError(f"打开Excel文件失败: {str(e)}")
    
    def close_workbook(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            logger.info("已关闭Excel文件")
    
    def get_sheet_names(self) -> List[str]:
        """获取所有工作表名称"""
        if not self.workbook:
            self.open_workbook()
        return self.workbook.sheetnames
    
    def _clean_value(self, value: Any, replacement: Any = None) -> Any:
        """清理单元格值"""
        if value is None:
            return replacement
        
        if isinstance(value, str):
            cleaned = value.strip()
            # 识别空值标记
            empty_markers = ['', 'none', 'null', 'nan', 'n/a', '-', '—', '—']
            if cleaned.lower() in empty_markers:
                return replacement
            # 尝试转换为整数（仅纯数字）
            if cleaned.isdigit() and all(c in '0123456789' for c in cleaned):
                return int(cleaned)
            # 尝试转换为浮点数
            try:
                return float(cleaned)
            except ValueError:
                pass
            return cleaned
        
        return value
    
    def read_headers(self, sheet: Worksheet, header_row: int = 1, property_row: int = 2) -> ExcelHeader:
        """
        读取表头信息
        
        :param sheet: 工作表对象
        :param header_row: 中文表头行号（从1开始）
        :param property_row: 英文属性名行号（从1开始）
        :return: ExcelHeader对象
        """
        chinese_names = []
        english_names = []
        
        for col in range(1, sheet.max_column + 1):
            chinese_val = sheet.cell(row=header_row, column=col).value
            english_val = sheet.cell(row=property_row, column=col).value
            
            chinese_names.append(self._clean_value(chinese_val))
            english_names.append(self._clean_value(english_val))
        
        logger.debug(f"读取表头: 中文列名={chinese_names}, 英文列名={english_names}")
        return ExcelHeader(chinese_names, english_names)
    
    def read_sheet_data(
        self,
        sheet_name: str,
        header_row: int = 1,
        property_row: int = 2,
        data_start_row: int = 3,
        empty_value_replacement: Any = None,
        filter_func: Optional[callable] = None
    ) -> List[Dict[str, Any]]:
        """
        读取指定工作表的数据
        
        :param sheet_name: 工作表名称
        :param header_row: 中文表头行号
        :param property_row: 英文属性名行号
        :param data_start_row: 数据起始行号
        :param empty_value_replacement: 空值替换值
        :param filter_func: 数据过滤函数
        :return: 数据列表
        """
        if not self.workbook:
            self.open_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        
        sheet = self.workbook[sheet_name]
        header = self.read_headers(sheet, header_row, property_row)
        
        data = []
        row_count = 0
        skipped_count = 0
        
        for row_num in range(data_start_row, sheet.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_idx, english_name in enumerate(header.english_names):
                if not english_name:
                    continue
                
                cell_value = sheet.cell(row=row_num, column=col_idx + 1).value
                cleaned_value = self._clean_value(cell_value, empty_value_replacement)
                row_data[english_name] = cleaned_value
                
                if cleaned_value is not None:
                    has_data = True
            
            # 跳过空行
            if not has_data:
                skipped_count += 1
                continue
            
            # 应用过滤函数
            if filter_func and not filter_func(row_data):
                skipped_count += 1
                continue
            
            data.append(row_data)
            row_count += 1
        
        logger.info(f"工作表 '{sheet_name}' 读取完成: 总行数={len(data)}, 跳过={skipped_count}")
        return data
    
    def read_all_sheets(self) -> Dict[str, List[Dict[str, Any]]]:
        """读取所有工作表数据"""
        if not self.workbook:
            self.open_workbook()
        
        result = {}
        for sheet_name in self.workbook.sheetnames:
            if sheet_name and sheet_name[0].isascii():
                try:
                    data = self.read_sheet_data(sheet_name)
                    result[sheet_name] = data
                except Exception as e:
                    logger.error(f"读取工作表 '{sheet_name}' 失败: {str(e)}")
        
        return result
    
    def get_sheet_info(self, sheet_name: str) -> Dict[str, Any]:
        """获取工作表信息"""
        if not self.workbook:
            self.open_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        
        sheet = self.workbook[sheet_name]
        header = self.read_headers(sheet)
        
        return {
            'sheet_name': sheet_name,
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'chinese_headers': header.chinese_names,
            'english_headers': header.english_names
        }
    
    def __enter__(self):
        """上下文管理器入口"""
        self.open_workbook()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close_workbook()


def print_excel_info(file_path: str):
    """打印Excel文件信息"""
    try:
        with ExcelReader(file_path) as reader:
            print("=" * 60)
            print(f"Excel文件: {file_path}")
            print("=" * 60)
            print(f"工作表数量: {len(reader.get_sheet_names())}")
            print(f"工作表列表: {reader.get_sheet_names()}")
            print("-" * 60)
            
            for sheet_name in reader.get_sheet_names():
                info = reader.get_sheet_info(sheet_name)
                print(f"\n工作表: '{sheet_name}'")
                print(f"  行数: {info['max_row']}, 列数: {info['max_column']}")
                print(f"  中文列名: {info['chinese_headers']}")
                print(f"  英文列名: {info['english_headers']}")
            
            print("\n" + "=" * 60)
    
    except Exception as e:
        logger.error(f"读取Excel文件信息失败: {str(e)}")
        print(f"❌ 读取失败: {str(e)}")


if __name__ == "__main__":
    import argparse
    
    logging.basicConfig(level=logging.INFO)
    
    parser = argparse.ArgumentParser(description="Excel读取工具")
    parser.add_argument("input", help="输入Excel文件路径")
    parser.add_argument("-l", "--list", action="store_true", help="列出工作表信息")
    parser.add_argument("-s", "--sheet", help="指定工作表名称")
    
    args = parser.parse_args()
    
    if args.list:
        print_excel_info(args.input)
    elif args.sheet:
        with ExcelReader(args.input) as reader:
            data = reader.read_sheet_data(args.sheet)
            print(f"读取到 {len(data)} 行数据")
            if data:
                print("第一行数据:")
                print(data[0])
    else:
        print("请指定 -l 或 -s 参数")