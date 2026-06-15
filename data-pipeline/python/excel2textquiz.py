#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel数据转Text-Quiz模式JSON脚本

功能：
1. 读取level3_adaptive_quiz工作表
2. 按text_id分组
3. 生成text+quiz配对的JSON格式
4. correct_answer字母转数字索引
"""

import os
import json
import openpyxl

# 输入文件路径
INPUT_FILE = r"e:\cpp_discipline\wenyan-platform\数据\开发需求填写.dbt.xlsx"

# 输出目录
OUTPUT_DIR = r"e:\cpp_discipline\wenyan-platform\src\data\text-quiz"


def read_level3_data_directly(file_path):
    """直接按列位置读取level3_adaptive_quiz数据，处理表头错误"""
    print(f"📥 正在读取Excel文件: {file_path}")
    
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["level3_adaptive_quiz"]
    
    # 打印表头信息（用于调试）
    headers_row1 = []  # 中文列名
    headers_row2 = []  # 英文列名
    for col in range(1, sheet.max_column + 1):
        headers_row1.append(sheet.cell(row=1, column=col).value)
        headers_row2.append(sheet.cell(row=2, column=col).value)
    
    print(f"📋 中文列名: {headers_row1}")
    print(f"🔤 英文列名: {headers_row2}")
    
    data = []
    for row in range(3, sheet.max_row + 1):
        # 按列位置直接读取
        item = {
            "text_id": sheet.cell(row=row, column=1).value,
            "scenario_text": sheet.cell(row=row, column=2).value,
            "question_id": sheet.cell(row=row, column=3).value,
            "question_type": sheet.cell(row=row, column=4).value,
            "question_text": sheet.cell(row=row, column=5).value,
            "option_a": sheet.cell(row=row, column=6).value,  # F列表头错误写为question_text
            "option_b": sheet.cell(row=row, column=7).value,
            "option_c": sheet.cell(row=row, column=8).value,
            "option_d": sheet.cell(row=row, column=9).value,
            "correct_answer": sheet.cell(row=row, column=10).value,
            "explanation": sheet.cell(row=row, column=11).value,
            "difficulty": sheet.cell(row=row, column=12).value
        }
        
        # 处理空值
        for key in item:
            if item[key] is None or item[key] == '' or str(item[key]).strip() == '':
                item[key] = None
            else:
                item[key] = str(item[key]).strip()
        
        # 跳过text_id为空的行
        if item["text_id"]:
            data.append(item)
    
    print(f"📈 读取到 {len(data)} 行有效数据")
    
    return data


def letter_to_index(letter):
    """将字母(A/B/C/D)转换为数字索引(0/1/2/3)"""
    if not letter:
        return None
    letter = str(letter).strip().upper()
    if letter in ['A', 'B', 'C', 'D']:
        return ord(letter) - ord('A')
    return None


def generate_text_quiz(data):
    """按text_id分组生成Text-Quiz模式JSON"""
    groups = {}
    
    for row in data:
        text_id = row.get("text_id")
        if not text_id:
            continue
        
        if text_id not in groups:
            groups[text_id] = []
        
        # 构建quiz对象
        quiz = {
            "question_id": row.get("question_id"),
            "question_type": row.get("question_type") or "radio",
            "question_text": row.get("question_text"),
            "options": [
                row.get("option_a"),
                row.get("option_b"),
                row.get("option_c"),
                row.get("option_d")
            ],
            "correct_answer": letter_to_index(row.get("correct_answer")),
            "explanation": row.get("explanation"),
            "difficulty": row.get("difficulty")
        }
        
        # 构建item
        item = {
            "text": row.get("scenario_text"),
            "quiz": quiz
        }
        
        groups[text_id].append(item)
    
    return groups


def save_json(groups):
    """保存每个text_id对应的JSON文件"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    for text_id, items in groups.items():
        output = {
            "pageId": text_id,
            "items": items
        }
        
        output_path = os.path.join(OUTPUT_DIR, f"{text_id}.json")
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        
        # 输出验证信息
        print(f"\n✅ 已生成: {output_path}")
        print(f"   📊 总items数: {len(items)}")
        if items:
            first_item = items[0]
            print(f"   🔖 第一个item包含text: {bool(first_item.get('text'))}")
            print(f"   📝 第一个quiz选项数: {len([o for o in first_item['quiz']['options'] if o])}")
            print(f"   🎯 第一个quiz正确答案索引: {first_item['quiz']['correct_answer']}")


def main():
    print("=" * 60)
    print("Excel转Text-Quiz模式JSON工具")
    print("=" * 60)
    print(f"📥 输入文件: {INPUT_FILE}")
    print(f"📤 输出目录: {OUTPUT_DIR}")
    print("-" * 60)
    
    # 检查输入文件
    if not os.path.exists(INPUT_FILE):
        print(f"❌ 错误: 未找到文件 {INPUT_FILE}")
        return
    
    # 读取数据（使用直接按列读取的方法）
    data = read_level3_data_directly(INPUT_FILE)
    
    # 生成text-quiz数据
    groups = generate_text_quiz(data)
    
    # 保存JSON文件
    save_json(groups)
    
    print("\n" + "=" * 60)
    print("🎉 Text-Quiz模式JSON生成完成！")
    print("=" * 60)
    print(f"📦 共生成 {len(groups)} 个页面配置文件")
    
    # 验证question_id与text_id的一致性
    print("\n🔍 数据验证:")
    for text_id, items in groups.items():
        for item in items:
            q_id = item['quiz']['question_id']
            if q_id and text_id not in str(q_id):
                print(f"   ⚠️  警告: question_id '{q_id}' 与 text_id '{text_id}' 不一致")


if __name__ == "__main__":
    main()