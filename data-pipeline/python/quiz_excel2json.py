#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel数据转JSON脚本 - Quiz数据处理
处理level1_quiz, level2_dialog_quiz, level3_adaptive_quiz三个sheet
只读取WEN_01的数据
"""

import os
import json
import openpyxl

# 输入文件路径
INPUT_FILE = r"e:\cpp_discipline\wenyan-platform\数据\开发需求填写.dbt.xlsx"

# 输出目录
OUTPUT_DIR = r"e:\cpp_discipline\wenyan-platform\数据"


def read_excel_sheet_with_columns(file_path, sheet_name):
    """读取Excel指定sheet的数据，返回带列索引的数据"""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    
    # 获取表头（第2行是英文属性名）
    headers = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=2, column=col).value
        if cell_value:
            headers.append(str(cell_value).strip())
        else:
            headers.append(None)
    
    # 读取数据行（从第3行开始）
    data = []
    for row in range(3, sheet.max_row + 1):
        row_data = {}
        text_id = sheet.cell(row=row, column=1).value
        # 只处理WEN_01的数据
        if text_id != 'WEN_01':
            continue
        
        for col in range(1, sheet.max_column + 1):
            header = headers[col - 1]
            if header:
                cell_value = sheet.cell(row=row, column=col).value
                # 处理空值
                if cell_value is None or cell_value == '':
                    row_data[header] = None
                else:
                    row_data[header] = str(cell_value).strip()
        
        if row_data:
            data.append(row_data)
    
    return data


def generate_level1_quiz(data):
    """生成level1_quiz.json"""
    output = []
    for row in data:
        item = {
            "text_id": row.get("text_id"),
            "question_number": int(row.get("question_number")) if row.get("question_number") else None,
            "question_text": row.get("question_text"),
            "option_a": row.get("option_a"),
            "option_b": row.get("option_b"),
            "option_c": row.get("option_c"),
            "option_d": row.get("option_d"),
            "audio_file": row.get("audio_file"),
            "difficulty": row.get("difficulty"),
            "correct_answer": row.get("correct_answer"),
            "explanation": row.get("explanation"),
            "question_type": row.get("question_type")
        }
        output.append(item)
    return output


def generate_level2_dialog(data):
    """生成level2_dialog.json"""
    output = []
    for row in data:
        # 只有pre_dialog非空时才生成dialog记录
        if row.get("pre_dialog"):
            item = {
                "text_id": row.get("text_id"),
                "pre_dialog": row.get("pre_dialog"),
                "audio_file": row.get("audio_file"),
                "icon_dialog": row.get("icon_dialog")
            }
            output.append(item)
    return output


def generate_level2_quiz(data):
    """生成level2_quiz.json"""
    output = []
    for row in data:
        # 只有question_text非空时才生成quiz记录
        if row.get("question_text"):
            item = {
                "text_id": row.get("text_id"),
                "question_number": int(row.get("question_number")) if row.get("question_number") else None,
                "question_text": row.get("question_text"),
                "option_a": row.get("option_a"),
                "option_b": row.get("option_b"),
                "option_c": row.get("option_c"),
                "option_d": row.get("option_d"),
                "audio_file": row.get("audio_file"),
                "difficulty": row.get("difficulty"),
                "correct_answer": row.get("correct_answer"),
                "explanation": row.get("explanation"),
                "question_type": row.get("question_type")
            }
            output.append(item)
    return output


def read_level3_adaptive_quiz_directly(file_path):
    """直接按列位置读取level3_adaptive_quiz，处理表头错误"""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["level3_adaptive_quiz"]
    
    data = []
    for row in range(3, sheet.max_row + 1):
        text_id = sheet.cell(row=row, column=1).value
        # 只处理WEN_01的数据
        if text_id != 'WEN_01':
            continue
        
        # 按列位置直接读取，忽略错误的表头
        item = {
            "text_id": text_id,
            "scenario_text": sheet.cell(row=row, column=2).value,
            "question_id": sheet.cell(row=row, column=3).value,
            "question_type": sheet.cell(row=row, column=4).value,
            "question_text": sheet.cell(row=row, column=5).value,
            "option_a": sheet.cell(row=row, column=6).value,  # F列表头错误写为question_text
            "option_b": sheet.cell(row=row, column=7).value,
            "option_c": sheet.cell(row=row, column=8).value,
            "option_d": sheet.cell(row=row, column=9).value,
            "correct_answer": sheet.cell(row=row, column=10).value,
            "explanation": sheet.cell(row=row, column=11).value,
            "difficulty": sheet.cell(row=row, column=12).value
        }
        
        # 处理空值
        for key in item:
            if item[key] is None or item[key] == '':
                item[key] = None
            else:
                item[key] = str(item[key]).strip()
        
        data.append(item)
    
    return data


def generate_level3_quiz(data):
    """生成level3_adaptive_quiz.json"""
    output = []
    for row in data:
        q_id = row.get("question_id")
        q_num = None
        if q_id:
            parts = str(q_id).split("_Q")
            if len(parts) > 1:
                q_num = int(parts[1])
        
        item = {
            "text_id": row.get("text_id"),
            "question_number": q_num,
            "question_text": row.get("question_text"),
            "option_a": row.get("option_a"),
            "option_b": row.get("option_b"),
            "option_c": row.get("option_c"),
            "option_d": row.get("option_d"),
            "audio_file": row.get("audio_file"),
            "difficulty": row.get("difficulty"),
            "correct_answer": row.get("correct_answer"),
            "explanation": row.get("explanation"),
            "question_type": row.get("question_type")
        }
        output.append(item)
    return output


def generate_level3_scenario_text(data):
    """生成level3_scenario_text.json"""
    output = []
    for row in data:
        if row.get("scenario_text"):
            q_id = row.get("question_id")
            q_num = None
            if q_id:
                parts = str(q_id).split("_Q")
                if len(parts) > 1:
                    q_num = int(parts[1])
            
            item = {
                "text_id": row.get("text_id"),
                "scenario_text": row.get("scenario_text"),
                "question_number": q_num
            }
            output.append(item)
    return output


def save_json(data, filename):
    """保存JSON文件"""
    output_path = os.path.join(OUTPUT_DIR, filename)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已生成: {output_path}")


def main():
    print("开始处理Excel数据...")
    
    # 处理level1_quiz
    print("\n处理level1_quiz...")
    level1_data = read_excel_sheet_with_columns(INPUT_FILE, "level1_quiz")
    level1_quiz = generate_level1_quiz(level1_data)
    save_json(level1_quiz, "level1_quiz.json")
    
    # 处理level2_dialog_quiz
    print("\n处理level2_dialog_quiz...")
    level2_data = read_excel_sheet_with_columns(INPUT_FILE, "level2_dialog_quiz")
    level2_dialog = generate_level2_dialog(level2_data)
    save_json(level2_dialog, "level2_dialog.json")
    level2_quiz = generate_level2_quiz(level2_data)
    save_json(level2_quiz, "level2_quiz.json")
    
    # 处理level3_adaptive_quiz（使用特殊处理）
    print("\n处理level3_adaptive_quiz...")
    level3_data = read_level3_adaptive_quiz_directly(INPUT_FILE)
    level3_quiz = generate_level3_quiz(level3_data)
    save_json(level3_quiz, "level3_adaptive_quiz.json")
    level3_scenario = generate_level3_scenario_text(level3_data)
    save_json(level3_scenario, "level3_scenario_text.json")
    
    print("\n✓ 所有JSON文件已生成完成！")


if __name__ == "__main__":
    main()